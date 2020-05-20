from openpyxl import load_workbook

class Cases:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
    def get_data(self): #获取excel的值
        wb=load_workbook(self.file_name) #打开表格文件
        sheet=wb[self.sheet_name]#打开sheet表单
        cases_data=[] #定义大列表
        for i in range(2,sheet.max_row+1): #循环获取单元格的值
            case = Cases()  # 初始化Cases类对象
            case.case_id=sheet.cell(row=i,column=1).value
            case.title=sheet.cell(row=i,column=2).value
            case.url=sheet.cell(row=i,column=3).value
            case.data=sheet.cell(row=i,column=4).value
            case.method=sheet.cell(row=i,column=5).value
            case.expected=sheet.cell(row=i,column=6).value
            cases_data.append(case) #将值追加至大列表中
        return cases_data #返回大列表的值
    def write_data(self,row,col,value): #写入值
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        sheet.cell(row,col).value=value
        wb.save(self.file_name) #保存表格

if __name__ == '__main__':
    from common import contant

    do_excel= DoExcel(contant.file_dir,'register').get_data()
    print(do_excel)
